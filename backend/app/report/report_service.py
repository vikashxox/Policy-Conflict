from __future__ import annotations

import io
import json
import re
from typing import Any
from datetime import datetime

# ReportLab imports
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing, Rect, String, Line

# openpyxl imports
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def safe_get(obj: Any, key: str, default: Any = None) -> Any:
    """Safely retrieves a key/attribute from a dict or object, supporting camelCase to snake_case translation."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        # Dict lookup
        val = obj.get(key)
        if val is not None:
            return val
        # Try snake_case in dict
        snake_key = re.sub(r'(?<!^)(?=[A-Z])', '_', key).lower()
        if snake_key != key:
            val = obj.get(snake_key)
            if val is not None:
                return val
        return default

    # Object attribute lookup
    val = getattr(obj, key, None)
    if val is not None:
        return val
    # Try snake_case attribute
    snake_key = re.sub(r'(?<!^)(?=[A-Z])', '_', key).lower()
    if snake_key != key:
        val = getattr(obj, snake_key, None)
        if val is not None:
            return val
    return default


class ReportService:
    def generate_report(self, title: str, payload: dict[str, Any], report_format: str = "json") -> dict[str, Any]:
        """Legacy helper kept for test backward compatibility."""
        content = json.dumps({"title": title, **payload}, indent=2)
        return {
            "title": title,
            "format": report_format,
            "content": content,
            "download_url": f"/reports/{title.lower().replace(' ', '-')}.{report_format}",
        }

    def generate_report_file(
        self,
        report_type: str,
        report_format: str,
        insights: dict[str, Any],
        findings: list[Any],
        policies: list[Any]
    ) -> tuple[bytes, str, str]:
        """
        Generates the report file based on the type and format.
        Returns: (file_bytes, content_type, filename)
        """
        report_type = report_type.lower().strip()
        report_format = report_format.lower().strip()

        # Normalize format name for extension/filename
        ext = "xlsx" if report_format == "excel" else report_format
        filename = f"{report_type}-report.{ext}"

        if report_format == "json":
            content_type = "application/json"
            file_bytes = self._generate_json(report_type, insights, findings, policies)
        elif report_format == "excel" or ext == "xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_bytes = self._generate_excel(report_type, insights, findings, policies)
        elif report_format == "pdf":
            content_type = "application/pdf"
            file_bytes = self._generate_pdf(report_type, insights, findings, policies)
        else:
            raise ValueError(f"Unsupported report format: {report_format}")

        return file_bytes, content_type, filename

    # ------------------------------------------------------------------
    # JSON Generation
    # ------------------------------------------------------------------
    def _generate_json(
        self,
        report_type: str,
        insights: dict[str, Any],
        findings: list[Any],
        policies: list[Any]
    ) -> bytes:
        data = self._build_report_data_structure(report_type, insights, findings, policies)
        return json.dumps(data, indent=2, default=str).encode("utf-8")

    # ------------------------------------------------------------------
    # Excel Generation
    # ------------------------------------------------------------------
    def _generate_excel(
        self,
        report_type: str,
        insights: dict[str, Any],
        findings: list[Any],
        policies: list[Any]
    ) -> bytes:
        wb = openpyxl.Workbook()
        # Remove default sheet
        default_sheet = wb.active
        if default_sheet is not None:
            wb.remove(default_sheet)

        # Style Definitions
        font_title = Font(name="Arial", size=16, bold=True, color="1E3A8A")
        font_subtitle = Font(name="Arial", size=10, italic=True, color="475569")
        font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Arial", size=10, bold=True, color="1E293B")
        font_normal = Font(name="Arial", size=10, color="334155")
        
        fill_primary = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        fill_healthy = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fill_warning = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
        fill_critical = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

        thin_side = Side(border_style="thin", color="CBD5E1")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # Sheet 1: Summary Page
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.views.sheetView[0].showGridLines = True
        
        ws_summary.append([])
        ws_summary.cell(row=2, column=2, value=f"{report_type.upper()} REPORT — SUMMARY").font = font_title
        ws_summary.cell(row=3, column=2, value=f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = font_subtitle
        ws_summary.append([])
        ws_summary.append([])

        # KPIs block
        ws_summary.cell(row=6, column=2, value="Key Performance Indicators").font = font_bold
        kpis = [
            ("Metric", "Value", "Status"),
            ("Overall Health Score", f"{insights['health_score']}%", insights['status'].upper()),
            ("Total Policies Analysed", insights['total_policies'], "ACTIVE"),
            ("Healthy Policies", insights.get('healthy_policies', 0), "HEALTHY"),
            ("Direct Policy Conflicts", insights['critical_findings'], "CRITICAL" if insights['critical_findings'] > 0 else "OK"),
            ("Stale Policies Detected", insights['stale_policy_count'], "WARNING" if insights['stale_policy_count'] > 0 else "OK"),
            ("Duplicate Policies", insights['duplicate_count'], "WARNING" if insights['duplicate_count'] > 0 else "OK"),
            ("Affected Policies", insights.get('affected_policies', 0), "ATTENTION REQUIRED" if insights.get('affected_policies', 0) > 0 else "OK")
        ]

        for i, row_data in enumerate(kpis):
            r = 7 + i
            for c, val in enumerate(row_data):
                cell = ws_summary.cell(row=r, column=2 + c, value=val)
                cell.font = font_header if i == 0 else font_normal
                cell.border = border_all
                if i == 0:
                    cell.fill = fill_primary
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left
                    if c == 2:
                        cell.alignment = align_center
                        if val in ("CRITICAL", "HIGH", "RED"):
                            cell.fill = fill_critical
                        elif val in ("WARNING", "MEDIUM", "YELLOW", "ATTENTION REQUIRED"):
                            cell.fill = fill_warning
                        elif val in ("OK", "HEALTHY", "GREEN", "ACTIVE"):
                            cell.fill = fill_healthy

        # Recommendations list
        r_start = 17
        ws_summary.cell(row=r_start, column=2, value="Key Recommendations").font = font_bold
        ws_summary.cell(row=r_start + 1, column=2, value="Priority").font = font_header
        ws_summary.cell(row=r_start + 1, column=2).fill = fill_primary
        ws_summary.cell(row=r_start + 1, column=3, value="Actionable Recommendation").font = font_header
        ws_summary.cell(row=r_start + 1, column=3).fill = fill_primary

        recs = insights.get("recommendations", [])
        if not recs:
            recs = ["No recommendations — policies are currently aligned."]
        for idx, rec in enumerate(recs):
            row_idx = r_start + 2 + idx
            cell_priority = ws_summary.cell(row=row_idx, column=2, value=f"P{idx+1}")
            cell_priority.font = font_bold
            cell_priority.alignment = align_center
            cell_priority.border = border_all

            cell_text = ws_summary.cell(row=row_idx, column=3, value=rec)
            cell_text.font = font_normal
            cell_text.alignment = align_left
            cell_text.border = border_all

        # Sheet 2: Detailed Findings
        ws_findings = wb.create_sheet(title="Findings Detail")
        ws_findings.views.sheetView[0].showGridLines = True
        ws_findings.append(["Detailed Findings and Conflicts"])
        ws_findings.cell(row=1, column=1).font = font_title
        ws_findings.append([])

        headers = ["ID", "Finding Type", "Severity", "Policy A", "Policy B", "Category", "Confidence", "Description", "Recommendation"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_findings.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = border_all

        for row_idx, f in enumerate(findings, start=4):
            # Resolve object vs dict attributes safely using safe_get
            f_id = safe_get(f, "id", row_idx - 3)
            f_type = safe_get(f, "finding_type", safe_get(f, "type", "Unknown"))
            f_severity = safe_get(f, "severity", "warning")
            f_policy_a = safe_get(f, "policy_a", safe_get(f, "policyA", "—"))
            f_policy_b = safe_get(f, "policy_b", safe_get(f, "policyB", "—"))
            f_category = safe_get(f, "category", "General")
            f_confidence = safe_get(f, "confidence", 100)
            f_desc = safe_get(f, "description", "")
            f_rec = safe_get(f, "recommendation", "")

            vals = [f"F-{f_id}", f_type, f_severity.upper(), f_policy_a, f_policy_b, f_category, f"{f_confidence}%", f_desc, f_rec]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws_findings.cell(row=row_idx, column=col_idx, value=v)
                cell.font = font_normal
                cell.border = border_all
                cell.alignment = align_wrap if col_idx in (8, 9) else align_left

                if col_idx == 3: # Severity coloring
                    cell.alignment = align_center
                    if v == "CRITICAL":
                        cell.fill = fill_critical
                    elif v == "WARNING":
                        cell.fill = fill_warning
                    else:
                        cell.fill = fill_healthy
                elif row_idx % 2 == 1 and col_idx != 3:
                    cell.fill = fill_zebra

        # Sheet 3: Policies Scope
        ws_policies = wb.create_sheet(title="Policy Governance")
        ws_policies.views.sheetView[0].showGridLines = True
        ws_policies.append(["Policies Covered by Analysis"])
        ws_policies.cell(row=1, column=1).font = font_title
        ws_policies.append([])

        p_headers = ["Policy Name", "Category", "Owner", "Department", "Version", "Last Reviewed", "Health Score", "Status"]
        for col_idx, h in enumerate(p_headers, start=1):
            cell = ws_policies.cell(row=3, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_primary
            cell.alignment = align_center
            cell.border = border_all

        for row_idx, p in enumerate(policies, start=4):
            p_name = safe_get(p, "name", "Unknown")
            p_cat = safe_get(p, "category", "General")
            p_owner = safe_get(p, "owner", "—")
            p_dept = safe_get(p, "department", "—")
            p_ver = safe_get(p, "version", "—")
            p_reviewed = safe_get(p, "last_reviewed", "—")
            p_health = safe_get(p, "health", 100)
            p_sev = safe_get(p, "severity", "healthy")

            vals = [p_name, p_cat, p_owner, p_dept, p_ver, p_reviewed, f"{p_health}%", p_sev.upper()]
            for col_idx, v in enumerate(vals, start=1):
                cell = ws_policies.cell(row=row_idx, column=col_idx, value=v)
                cell.font = font_normal
                cell.border = border_all
                cell.alignment = align_left

                if col_idx == 8: # Status coloring
                    cell.alignment = align_center
                    if v == "CRITICAL":
                        cell.fill = fill_critical
                    elif v == "WARNING":
                        cell.fill = fill_warning
                    else:
                        cell.fill = fill_healthy
                elif row_idx % 2 == 1 and col_idx != 8:
                    cell.fill = fill_zebra

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                # Skip title rows or long description wraps to avoid super wide columns
                for cell in col:
                    if cell.row in (1, 2) or (sheet.title == "Findings Detail" and cell.column in (8, 9)):
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)

            # Special wide wraps for descriptions
            if sheet.title == "Findings Detail":
                sheet.column_dimensions['H'].width = 45
                sheet.column_dimensions['I'].width = 45

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # PDF Generation (ReportLab)
    # ------------------------------------------------------------------
    def _generate_pdf(
        self,
        report_type: str,
        insights: dict[str, Any],
        findings: list[Any],
        policies: list[Any]
    ) -> bytes:
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf,
            pagesize=letter,
            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
        )

        styles = getSampleStyleSheet()
        
        # Color Palette
        c_primary = colors.HexColor("#1e3a8a")
        c_secondary = colors.HexColor("#334155")
        c_accent = colors.HexColor("#ef4444")
        c_neutral_dark = colors.HexColor("#0f172a")
        c_neutral_light = colors.HexColor("#f8fafc")
        c_border = colors.HexColor("#cbd5e1")

        # Paragraph Styles
        style_title = ParagraphStyle(
            "DocTitle",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=24,
            leading=28,
            textColor=c_primary,
            spaceAfter=6
        )
        style_subtitle = ParagraphStyle(
            "DocSubtitle",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=c_secondary,
            spaceAfter=20
        )
        style_h1 = ParagraphStyle(
            "H1",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=14,
            leading=18,
            textColor=c_primary,
            spaceBefore=14,
            spaceAfter=8,
            keepWithNext=True
        )
        style_body = ParagraphStyle(
            "Body",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13.5,
            textColor=c_neutral_dark,
            spaceAfter=8
        )
        style_bold = ParagraphStyle(
            "BodyBold",
            parent=style_body,
            fontName="Helvetica-Bold"
        )
        style_table_header = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            leading=11,
            textColor=colors.white,
            alignment=1 # Center
        )
        style_table_cell = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=c_neutral_dark
        )

        story = []

        # Header Block
        story.append(Paragraph(f"{report_type.capitalize()} Report", style_title))
        story.append(Paragraph(f"Policy Guardian Governance System  |  Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", style_subtitle))
        story.append(Spacer(1, 10))

        # Visual Decorative Horizontal Bar
        d_bar = Drawing(532, 4)
        d_bar.add(Rect(0, 0, 532, 4, fillColor=c_primary, strokeColor=None))
        story.append(d_bar)
        story.append(Spacer(1, 15))

        # 1. Executive Summary & KPIs
        story.append(Paragraph("1. Executive Summary", style_h1))
        story.append(Paragraph(insights.get("org_risk_summary", ""), style_body))
        story.append(Spacer(1, 10))

        # Map 'critical'/'warning'/'healthy' to valid ReportLab color tags
        raw_status = insights.get("status", "healthy").lower()
        status_color = "red" if raw_status == "critical" else ("orange" if raw_status == "warning" else "green")

        # KPI block drawing / table
        kpi_data = [
            [
                Paragraph("<b>Health Score</b>", style_body),
                Paragraph(f"<font color='{status_color}'><b>{insights['health_score']}%</b> ({raw_status.upper()})</font>", style_body),
                Paragraph("<b>Total Policies</b>", style_body),
                Paragraph(str(insights['total_policies']), style_body)
            ],
            [
                Paragraph("<b>Direct Conflicts</b>", style_body),
                Paragraph(f"<font color='red'><b>{insights['critical_findings']}</b></font>", style_body),
                Paragraph("<b>Stale Policies</b>", style_body),
                Paragraph(f"<b>{insights['stale_policy_count']}</b>", style_body)
            ]
        ]
        t_kpis = Table(kpi_data, colWidths=[120, 146, 120, 146])
        t_kpis.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), c_neutral_light),
            ('BOX', (0,0), (-1,-1), 1, c_border),
            ('INNERGRID', (0,0), (-1,-1), 0.5, c_border),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(t_kpis)
        story.append(Spacer(1, 15))

        # 2. Key Actionable Recommendations
        story.append(Paragraph("2. Prioritized Action Recommendations", style_h1))
        recs = insights.get("recommendations", [])
        if not recs:
            story.append(Paragraph("No critical issues found. Maintain current configuration and schedules.", style_body))
        else:
            for idx, r in enumerate(recs, start=1):
                bullet_text = f"<b>Priority {idx}:</b> {r}"
                story.append(Paragraph(bullet_text, style_body))
        story.append(Spacer(1, 15))

        # 3. Compliance Summary / Framework Coverage
        story.append(Paragraph("3. Compliance Mapping & Framework Status", style_h1))
        comp_summary = insights.get("compliance_summary", [])
        if not comp_summary:
            story.append(Paragraph("No active compliance frameworks mapped. Upload policies to populate framework details.", style_body))
        else:
            comp_table_data = [[
                Paragraph("Framework", style_table_header),
                Paragraph("Alignment Score", style_table_header),
                Paragraph("Affected Controls/Policies", style_table_header),
                Paragraph("Risk Level Status", style_table_header)
            ]]
            for c_item in comp_summary:
                status_color = "red" if c_item['status'] == "non-compliant" else ("orange" if c_item['status'] == "at-risk" else "green")
                comp_table_data.append([
                    Paragraph(f"<b>{c_item['framework']}</b>", style_table_cell),
                    Paragraph(f"{c_item['score']}%", style_table_cell),
                    Paragraph(str(c_item['affected']), style_table_cell),
                    Paragraph(f"<font color='{status_color}'><b>{c_item['status'].upper()}</b></font>", style_table_cell)
                ])

            t_comp = Table(comp_table_data, colWidths=[120, 120, 150, 142])
            t_comp.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), c_primary),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('GRID', (0,0), (-1,-1), 0.5, c_border),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, c_neutral_light])
            ]))
            story.append(t_comp)
        story.append(Spacer(1, 15))

        # 4. Detailed Findings (Direct Conflicts, Duplicates, and Staleness)
        story.append(Paragraph("4. Detailed Audit & Policy Findings", style_h1))
        if not findings:
            story.append(Paragraph("No active findings or policy conflicts detected in the system.", style_body))
        else:
            findings_table_data = [[
                Paragraph("Type / ID", style_table_header),
                Paragraph("Policy Context", style_table_header),
                Paragraph("Confidence", style_table_header),
                Paragraph("Finding Details & Resolution", style_table_header)
            ]]
            for idx, f in enumerate(findings[:12], start=1):
                f_id = safe_get(f, "id", idx)
                f_type = safe_get(f, "finding_type", safe_get(f, "type", "Unknown"))
                f_policy_a = safe_get(f, "policy_a", safe_get(f, "policyA", "—"))
                f_policy_b = safe_get(f, "policy_b", safe_get(f, "policyB", "—"))
                f_confidence = safe_get(f, "confidence", 100)
                f_desc = safe_get(f, "description", "")
                f_rec = safe_get(f, "recommendation", "")

                policy_context = f"{f_policy_a}"
                if f_policy_b and f_policy_b != "—":
                    policy_context += f"<br/><b>vs</b><br/>{f_policy_b}"

                finding_desc = f"<b>Desc:</b> {f_desc}<br/><b>Rec:</b> {f_rec}"

                findings_table_data.append([
                    Paragraph(f"<b>{f_type}</b><br/>F-{f_id}", style_table_cell),
                    Paragraph(policy_context, style_table_cell),
                    Paragraph(f"{f_confidence}%", style_table_cell),
                    Paragraph(finding_desc, style_table_cell)
                ])

            t_findings = Table(findings_table_data, colWidths=[110, 110, 60, 252])
            t_findings.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), c_primary),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('VALIGN', (0,0), (-1,-1), 'TOP'),
                ('GRID', (0,0), (-1,-1), 0.5, c_border),
                ('TOPPADDING', (0,0), (-1,-1), 6),
                ('BOTTOMPADDING', (0,0), (-1,-1), 6),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, c_neutral_light])
            ]))
            story.append(KeepTogether([t_findings]))

        if len(findings) > 12:
            story.append(Spacer(1, 8))
            story.append(Paragraph(f"<i>* Showing top 12 of {len(findings)} total findings. Please export as JSON or Excel for full audit log.</i>", style_body))

        # Build PDF
        doc.build(story)
        return buf.getvalue()

    # ------------------------------------------------------------------
    # Data Normalizer
    # ------------------------------------------------------------------
    def _build_report_data_structure(
        self,
        report_type: str,
        insights: dict[str, Any],
        findings: list[Any],
        policies: list[Any]
    ) -> dict[str, Any]:
        """Shape standard output payload dictionary to fit specific report type needs."""
        structured_findings = []
        for idx, f in enumerate(findings):
            f_id = safe_get(f, "id", idx + 1)
            f_type = safe_get(f, "finding_type", safe_get(f, "type", "Unknown"))
            f_severity = safe_get(f, "severity", "warning")
            f_policy_a = safe_get(f, "policy_a", safe_get(f, "policyA", "—"))
            f_policy_b = safe_get(f, "policy_b", safe_get(f, "policyB", "—"))
            f_category = safe_get(f, "category", "General")
            f_confidence = safe_get(f, "confidence", 100)
            f_desc = safe_get(f, "description", "")
            f_rec = safe_get(f, "recommendation", "")
            f_compliance = safe_get(f, "compliance", "Governance")
            f_status = safe_get(f, "status", "open")

            structured_findings.append({
                "id": f"F-{f_id}",
                "type": f_type,
                "severity": f_severity,
                "policyA": f_policy_a,
                "policyB": f_policy_b,
                "category": f_category,
                "confidence": f_confidence,
                "description": f_desc,
                "recommendation": f_rec,
                "compliance": f_compliance,
                "status": f_status
            })

        structured_policies = []
        for p in policies:
            structured_policies.append({
                "name": safe_get(p, "name", "Unknown"),
                "category": safe_get(p, "category", "General"),
                "owner": safe_get(p, "owner", "—"),
                "department": safe_get(p, "department", "—"),
                "version": safe_get(p, "version", "—"),
                "effectiveDate": safe_get(p, "effective_date", "—"),
                "lastReviewed": safe_get(p, "last_reviewed", "—"),
                "health": safe_get(p, "health", 100),
                "severity": safe_get(p, "severity", "healthy"),
                "status": safe_get(p, "status", "active")
            })

        return {
            "metadata": {
                "report_type": report_type.upper(),
                "generated_at": datetime.now().isoformat(),
                "system": "Policy Guardian GRC Platform"
            },
            "summary": {
                "health_score": insights["health_score"],
                "status": insights["status"],
                "description": insights["org_risk_summary"],
                "total_policies": insights["total_policies"],
                "affected_policies": insights.get("affected_policies", 0),
                "metrics": {
                    "conflicts": insights["critical_findings"],
                    "stale": insights["stale_policy_count"],
                    "duplicates": insights["duplicate_count"]
                }
            },
            "recommendations": insights.get("recommendations", []),
            "compliance_frameworks": insights.get("compliance_summary", []),
            "findings": structured_findings,
            "policies": structured_policies
        }
